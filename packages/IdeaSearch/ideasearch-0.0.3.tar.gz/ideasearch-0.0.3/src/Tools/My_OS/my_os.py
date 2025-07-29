import openpyxl
import os
import sys
import scipy.io
import numpy as np
import pandas as pd
import random


__all__ = [
    "guarantee_path_exist",
    "assert_path_exist",
    "append_to_file",
    "read_table_item", 
    "print_table",
    "merge_mat_files",
    "check_mat_part",
    "load_mat_part",
    "delete_lines_after",
    "new_table",
    "write_table",
    "write_table_item",
    "save_table",
]


def guarantee_path_exist(path):
    """
    确保给定路径存在，如果不存在则创建它。

    参数:
    path (str): 要检查或创建的路径。
    """

    # 检查路径是否存在
    if not os.path.exists(path):
        # 创建路径
        os.makedirs(path)
        
 
def assert_path_exist(path, err_message = None):
    """
    断言给定路径存在
    """

    # 检查路径是否存在
    if not os.path.exists(path):
        if err_message is None:
            assert False, f"程序出错，路径{path}不存在！"
        else:
            assert False, err_message
        

def append_to_file(file_path, content_str, end="\n"):
    """
    将指定内容附加到文件末尾。

    参数:
    file_path (str): 目标文件的路径。
    content_str (str): 要写入文件的内容。
    end (str, optional): 内容结尾的字符，默认为换行符。
    """

    # 以追加模式打开文件并写入内容
    with open(file_path, "a", encoding="utf-8") as file:
        file.write(content_str + end)
        

table_reader_max_cache_size = 1024 * 1024 * 1  # 1MB


class TableReader:
    """
    用于读取表格数据的类，支持缓存功能。

    属性:
    max_cache_size (int): 最大缓存容量（以字节为单位）。
    cache (dict): 存储已加载表格的缓存。
    current_cache_size (int): 当前缓存占用的大小。
    """

    def __init__(self, max_cache_size):
        """
        初始化 TableReader 实例。

        参数:
        max_cache_size (int): 最大缓存容量。
        """

        self.max_cache_size = max_cache_size  
        self.cache = {}
        self.current_cache_size = 0

    def _load_table(self, table_path):
        """
        加载指定路径的表格数据到缓存中。

        参数:
        table_path (str): 表格文件的路径。
        
        抛出:
        FileNotFoundError: 如果文件未找到。
        Exception: 如果读取表格时发生错误。
        """

        # 检查缓存中是否已存在该表格
        if table_path not in self.cache:
            # 检查文件是否存在
            if not os.path.exists(table_path):
                raise FileNotFoundError(f"文件未找到: {table_path}")

            try:
                # 加载工作簿和活动工作表
                workbook = openpyxl.load_workbook(table_path)
                sheet = workbook.active
                
                # 读取表格数据，第一行作为列名，第一列作为行名
                headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]
                row_names = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
                
                # 构建表格数据字典
                table_data = {
                    row_name: {header: sheet.cell(row=row, column=col).value 
                                for col, header in enumerate(headers, start=2)}
                    for row, row_name in enumerate(row_names, start=2)
                }
                
                # 计算当前表格数据的大小（简单估算）
                table_size = sys.getsizeof(table_data)

                # 更新缓存和当前缓存大小
                self.cache[table_path] = table_data
                self.current_cache_size += table_size

                # 确保缓存不超过最大容量
                while self.current_cache_size > self.max_cache_size:
                    oldest_table = next(iter(self.cache))
                    self.current_cache_size -= sys.getsizeof(self.cache[oldest_table])
                    del self.cache[oldest_table]  # 移除最早加载的表格
            except Exception as e:
                raise Exception(f"读取表格时发生错误: {e}")

    def _read_table(self, table_path):
        """
        读取指定路径的表格数据。

        参数:
        table_path (str): 表格文件的路径。
        
        返回:
        dict: 表格数据。
        """
        
        self._load_table(table_path)
        return self.cache[table_path]

    def _get_value(self, table_path, row_name, column_name):
        """
        获取指定表格中某一行和列的值。

        参数:
        table_path (str): 表格文件的路径。
        row_name (str): 行名。
        column_name (str): 列名。
        
        返回:
        any: 对应的单元格值，如果未找到则返回 None。
        """
        
        table_info = self._read_table(table_path)
        row_data = table_info.get(row_name)
        return row_data.get(column_name) if row_data else None

    def _print_table(self, table_path):
        """
        打印指定表格的所有行和列的数据。

        参数:
        table_path (str): 表格文件的路径。
        """
        
        table_info = self._read_table(table_path)
        for row_name, row_data in table_info.items():
            print(f"{row_name}: {row_data}")


# 创建 TableReader 实例
table_reader = TableReader(max_cache_size=table_reader_max_cache_size)


def read_table_item(table_path, row_name, column_name):
    """
    从指定表格中读取某一行和列的值。

    参数:
    table_path (str): 表格文件的路径。
    row_name (str): 行名。
    column_name (str): 列名。
    
    返回:
    any: 对应的单元格值。
    """
    
    return table_reader._get_value(table_path, row_name, column_name)


def print_table(table_path):
    """
    打印指定表格的所有行和列的数据。

    参数:
    table_path (str): 表格文件的路径。
    """
    
    return table_reader._print_table(table_path)


def merge_mat_files(input_mat_paths, output_mat_path):
    """
    将路径列表 `input_mat_paths` 中的所有 `.mat` 文件合并成一个 `.mat` 文件。

    参数:
    input_mat_paths (list of str): 包含所有待合并 `.mat` 文件路径的列表。
    output_mat_path (str): 合并后的 `.mat` 文件的输出路径。
    
    说明:
    - 如果遇到重复的属性（attribute），以最先处理的 `.mat` 文件中的值为准，忽略后续文件中相同的属性。
    - 合并后的数据将保存到 `output_mat_path` 指定的路径中。
    """

    # 检查输入路径的有效性
    for mat_path in input_mat_paths:
        if not os.path.exists(mat_path) or not mat_path.endswith('.mat'):
            print(f"无效的路径: {mat_path}，请检查该路径下是否存在一个 `.mat` 文件")
            return
        
    # 初始化一个空字典，用于存储合并后的数据
    merged_data = {}
    
    # 遍历每个 `.mat` 文件路径
    for mat_path in input_mat_paths:
        # 加载 `.mat` 文件中的数据
        data = scipy.io.loadmat(mat_path)
        for key, value in data.items():
            # 忽略IdeaSearcher生成的键（以 `__` 开头的键）
            if key.startswith('__'):
                continue
            # 如果键不存在于 `merged_data` 中，则将其添加到 `merged_data`
            if key not in merged_data:
                merged_data[key] = value
    
    # 将合并后的数据保存到目标 `.mat` 文件中
    scipy.io.savemat(output_mat_path, merged_data, do_compression=True)


def check_mat_part(mat_path, attribute):
    """
    检查给定的 `.mat` 文件路径是否有效，并且文件中包含指定的属性。

    参数:
    mat_path (str): `.mat` 文件的路径。
    attribute (str): 要检查的属性名。
    
    返回:
    int: 如果路径有效且文件中包含指定属性，返回 1；否则返回 0。
    """

    # 检查路径是否为文件
    if not os.path.isfile(mat_path):
        return 0
    
    try:
        # 获取文件中的所有变量
        variables = scipy.io.whosmat(mat_path)
        variable_names = [var[0] for var in variables]
        
        # 检查属性是否存在
        if attribute in variable_names:
            return 1
        return 0
    except Exception:
        return 0


# mat文件的cache
mat_buffer = dict()
mat_buffer_capacity = 1


def load_mat_part(mat_path, attribute):
    """
    从 `.mat` 文件中加载指定属性的 ndarray。

    参数:
    mat_path (str): `.mat` 文件的路径。
    attribute (str): 要加载的属性名。
    
    返回:
    numpy.ndarray: 指定属性的数据数组。
    
    抛出:
    AssertionError: 如果路径无效或文件中不包含指定属性。
    """

    # 检查缓存中是否已有数据
    if mat_path in mat_buffer:
        assert attribute in mat_buffer[mat_path], f"Error: mat_path {mat_path} invalid, or the file doesn't contain attribute {attribute}."
        return np.array(mat_buffer[mat_path][attribute])
    else:
        # 检查文件及属性有效性
        assert check_mat_part(mat_path, attribute), f"Error: mat_path {mat_path} invalid, or the file doesn't contain attribute {attribute}."
        
        # 加载 `.mat` 文件
        mat_data = scipy.io.loadmat(mat_path)
        
        # 管理缓存容量
        if len(mat_buffer) == mat_buffer_capacity:
            mat_buffer.pop(random.choice(list(mat_buffer.keys())))
        
        mat_buffer[mat_path] = mat_data
        return np.array(mat_data[attribute])
    

def delete_lines_after(file_path, line_number):  
    """
    删除文件中指定行号之后的所有行。

    参数:
    file_path (str): 要处理的文件路径。
    line_number (int): 指定行号，之后的行将被删除。
    """

    # 创建一个列表来存储要保留的行  
    lines_to_keep = []  
  
    # 尝试打开文件并读取内容  
    try:  
        with open(file_path, 'r', encoding='utf-8') as file:  
            # 读取文件的每一行，直到达到指定的行号  
            for i, line in enumerate(file, 1):  
                if i <= line_number:  
                    lines_to_keep.append(line)  
  
        # 覆盖原文件，或写入新文件  
        with open(file_path, 'w', encoding='utf-8') as file:  
            # 将要保留的行写回到文件  
            for line in lines_to_keep:  
                file.write(line)  
  
    except FileNotFoundError:  
        print(f"File {file_path} not found, will automatically create a new one!")  
    except Exception as e:  
        print(f"发生错误: {e}") 


def new_table(column_name_list):
    """
    创建一个新的表格，接收一个列名列表
    :param column_name_list: 表格的列名列表
    :return: 返回一个空的 DataFrame 对象
    """
    return pd.DataFrame(columns=column_name_list)


def write_table(table, row_index, column_index, value):
    """
    根据行列索引更新表格中的值，智能扩展 DataFrame。
    :param table: 表格对象（pandas DataFrame）
    :param row_index: 行索引
    :param column_index: 列索引
    :param value: 要写入的值
    :return: 更新后的表格
    """
    
    row_index = row_index - 1
    
    # 确保行和列索引在当前 DataFrame 范围内
    num_rows, num_columns = table.shape

    # 扩展行数
    if row_index >= num_rows:
        # 扩展行
        additional_rows = row_index - num_rows + 1
        # 向下扩展空行
        table = pd.concat([table, pd.DataFrame([[None]*num_columns] * additional_rows, columns=table.columns)], ignore_index=True)

    # 扩展列数
    if column_index >= num_columns:
        # 扩展列
        additional_columns = column_index - num_columns + 1
        # 向右扩展空列
        for i in range(additional_columns):
            table[table.columns[-1] + f'_new_{i}'] = [None] * table.shape[0]

    # 更新指定位置的值
    table.iloc[row_index, column_index] = value
    
    return table


def write_table_item(table, row_name, column_name, value):
    """
    根据行名（第一列）和列名（第一行）更新表格中的值
    :param table: 表格对象（pandas DataFrame）
    :param row_name: 行名（基于第一列）
    :param column_name: 列名（基于第一行）
    :param value: 要写入的值
    :return: 更新后的表格
    """
    # 检查行名是否在第一列
    if row_name not in table.iloc[:, 0].values:
        raise ValueError(f"Row name '{row_name}' not found in the first column.")

    # 检查列名是否在列名中
    if column_name not in table.columns:
        raise ValueError(f"Column name '{column_name}' not found in the table.")

    # 获取行索引（第一列的索引）
    row_index = table[table.iloc[:, 0] == row_name].index[0]
    
    # 获取列索引（列名的索引）
    column_index = table.columns.get_loc(column_name)

    # 更新指定位置的值
    table.iloc[row_index, column_index] = value
    
    return table




def save_table(table, table_path):
    """
    将表格保存到指定路径
    :param table: 表格对象（pandas DataFrame）
    :param table_path: 保存的文件路径
    """
    table.to_excel(table_path, index=False)



