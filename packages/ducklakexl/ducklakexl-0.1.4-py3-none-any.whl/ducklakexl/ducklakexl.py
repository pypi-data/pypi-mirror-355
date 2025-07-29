import duckdb
import pandas as pd
import numpy as np
import os
import asyncio
import string
import requests
import aiohttp
import openpyxl
import tempfile


class DuckLakeXL():
    """Wrap DuckDB ducklake functionality in a way that syncs a local copy of the 
    catalog to an Excel file (remote on SharePoint/OneDrive or local)"""

    def __init__(self,
                 excel_path: str,
                 data_path: str,
                 local_catalog: str = 'ducklakexl.ducklake',
                 duckdb_conn: duckdb.DuckDBPyConnection | None = None,
                 ducklake_name: str = 'my_ducklake',
                 #encrypted: bool = False,
                 custom_cert_store = 'certifi',
                 create_if_missing: bool = False,
                 folder_path: str | None = None,
                 drive_id: str | None = None,
                 read_shared_files: bool = False
                 ):
        """
        Initialize DuckLakeXL instance.

        Wrap DuckDB ducklake functionality in a way that syncs a local copy of the
        catalog to an Excel file (remote on SharePoint/OneDrive or local).

        Args:
            excel_path (str): The file name or path for the Excel workbook. For OneDrive/SharePoint, this should be the file name you want to see on OneDrive (not a local path). For local Excel, provide a local file path. If creating a new OneDrive workbook, this is the name that will appear on OneDrive.
            data_path (str): Path where the data lake files will be stored.
            local_catalog (str, optional): Name of the DuckLake catalog. Defaults to 'ducklakexl.ducklake'.
            duckdb_conn (duckdb.DuckDBPyConnection, optional): Existing DuckDB connection. If None, creates new in-memory connection. Defaults to None.
            ducklake_name (str, optional): Name for the attached DuckLake instance. Defaults to 'my_ducklake'.
            custom_cert_store (str, optional): Custom certificate store to use for SSL. Defaults to 'certifi'.
            create_if_missing (bool, optional): If True and using OneDrive, will create a new workbook with the name specified by excel_path in the given folder_path. Defaults to False.
            folder_path (str, optional): Folder path on OneDrive where the workbook should be created (relative to the root of the drive). Only used if create_if_missing is True.
            drive_id (str, optional): The OneDrive drive ID where the workbook is located or should be created. Must be specified for OneDrive/SharePoint mode.
            read_shared_files (bool, optional): If true, OneDrive API requests will use the files.readwrite.all scope, and have access to your files AND files shared with you. If false (default), it just uses files.readwrite scope, which can only access your files.

        Notes:
            - For OneDrive/SharePoint mode, you must specify both excel_path (the file name you want to see on OneDrive) and drive_id.
            - If drive_id is provided, OneDrive mode is used regardless of the excel_path extension.
            - For local Excel files, provide a local file path and do not specify drive_id.

        """
        self.excel_path = excel_path
        self.data_path = data_path
        self.local_catalog = local_catalog
        self.ducklake_name = ducklake_name
        #self.encrypted = encrypted
        self.custom_cert_store = custom_cert_store
        self.drive_id = drive_id
        self.create_if_missing = create_if_missing
        self.folder_path = folder_path
        self.read_shared_files = read_shared_files
        self._pick_client()

        # the MSGraph sdk is all async coroutines - for initial simplicity, keep this class all sync
        # whenever we call one of the graph functions, need to wrap in a loop:
        if asyncio.get_event_loop_policy()._local._loop is not None: # for the case this module is called from Jupyter or a script with its own loop; if calling script creates own loop, will need to be done before initializing ducklakeXL
            self.loop = asyncio.get_event_loop()
        else:
            self.loop = asyncio.new_event_loop()

        self._initialize_client()


        # if user supplies an existing duckdb connection, use that, otherwise we create a new in-memory one
        if duckdb_conn:
            self.db = duckdb_conn
        else:
            self.db = duckdb.connect()
        
        self._initialize_ducklake()


    def _pick_client(self):
        """
        Decide if we are using a local Excel file or OneDrive/SharePoint based on the parameters.

        If drive_id is specified, always use OneDrive/SharePoint mode, regardless of excel_path.
        Otherwise, use local Excel mode if excel_path ends with .xlsx and is not a URL.
        """
        if self.drive_id:
            self.client_type = 'onedrive'
        elif self.excel_path.endswith('.xlsx') and not self.excel_path.startswith('https://'):
            self.client_type = 'excel'
        else:
            self.client_type = 'onedrive'


    def _acquire_token(self,force_refresh=False):
        """retrieve existing token, if cached. otherwise request new
        Returns a headers dict to pass to request"""
        accounts = self.app.get_accounts(username=None)
        if len(accounts) > 0:
            account = accounts[0]  # Simulate user selection
        else:
            account = None
        result = self.app.acquire_token_silent(self.scopes, account=account)

        if (not result) or force_refresh:
            result = self.app.acquire_token_interactive(scopes=self.scopes)
            if "access_token" in result:
                print('Authentication successful.')
                access_token = result['access_token']
                # Use the access token to call Microsoft Graph API
                headers = {'Authorization': f'Bearer {access_token}'}
                response = requests.get('https://graph.microsoft.com/v1.0/me', headers=headers)

                if response.status_code == 200:
                    user_data = response.json()
                    self.username = user_data['userPrincipalName']
                else:
                    print(f"API call failed with status code {response.status_code}: {response.text}")
            else:
                print(f"Authentication failed: {result.get('error_description')}")
        else:
            #print("Successfully retrieved token from cache!")
            access_token = result['access_token']
            headers = {'Authorization': f'Bearer {access_token}'}

        return headers

    def _initialize_client(self):
        
        if self.client_type == 'onedrive':
            import msal
            import requests

            # get the MS EntraID App Client_id - for now assume in a .env file or defined as env var
            import importlib.util
            dotenv_spec = importlib.util.find_spec("dotenv")
            if dotenv_spec is not None:
                from dotenv import load_dotenv
                load_dotenv()

            CLIENT_ID = os.getenv('CLIENT_ID')
            AUTHORITY = f'https://login.microsoftonline.com/consumers'


            # Define your application (client) ID and the scopes required
            client_id = CLIENT_ID
            scopes = ['Files.ReadWrite', 'User.Read']  # Add other scopes as needed
            if self.read_shared_files:
                scopes.append('Files.ReadWrite.All')
            self.scopes = scopes
            self.username = None # initialize to None and update on first login - only used to maintain in-memory token cache

            # Create a public client application
            self.token_cache = msal.TokenCache() # The TokenCache() is in-memory.
            self.app = msal.PublicClientApplication(CLIENT_ID, 
                                               authority=AUTHORITY,
                                               token_cache=self.token_cache
                                               )

            initial_header = self._acquire_token() # call to get an initial token up-front
            
            #ensure we use the user's preferred ssl context
            import ssl
            if self.custom_cert_store == 'certifi':
                import certifi
                self.ssl_context = ssl.create_default_context(cafile=certifi.where())
            else:
                self.ssl_context = ssl.create_default_context()
            
            # check if file exists and create if needed
            self._initialize_onedrive_file()

        elif self.client_type == 'excel':
            pass # nothing needed here - may need separate client initialization for fsspec files

        else:
            raise NotImplementedError("Only OneDrive and Excel client types are currently supported")
        

    def _initialize_onedrive_file(self):
        # check if file already exists
        #if not, create it

        self.item_id = None
        if not self.drive_id:
            if '!' in self.excel_path:
                # excel path of form {drive_id}!{item-specific-part}
                self.drive_id = self.excel_path.split('!')[0]
                self.item_id = self.excel_path
            else:
                raise ValueError("drive_id must be specified for OneDrive mode")
        # Attempt to get the item ID of an existing file, if excel_path wasn't specified with item ID to begin with
        if not self.item_id:
            try:
                item_info = self.loop.run_until_complete(self._get_onedrive_item(self.excel_path,self.folder_path))
                self.item_id = item_info.get('id')
            except FileNotFoundError:
                if self.create_if_missing:
                    # Create workbook if missing
                    self._create_onedrive_workbook(self.excel_path, self.folder_path)
                    # _create_onedrive_workbook sets self.item_id
                else:
                    error_message = f"File '{self.excel_path}' not found on OneDrive and create_if_missing=False"
                    raise FileNotFoundError(error_message)
            

    async def _get_onedrive_item(self, fname: str, folder_path: str | None = None) -> dict:
        """Check if a file exists in OneDrive; return item info or raise FileNotFoundError"""
        headers = self._acquire_token()
        headers.update({})
        # URL-encode the file path
        #encoded_name = requests.utils.requote_uri(fname)
        if folder_path:
            url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{folder_path.rstrip('/').lstrip('/')}/{fname}"
        else:
            url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{fname}"
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            async with session.get(url, headers=headers) as resp:
                if resp.status == 200:
                    return await resp.json(content_type=None)
                elif resp.status == 404:
                    raise FileNotFoundError(f"Item '{fname}' not found on OneDrive")
                else:
                    resp.raise_for_status()
        

    def _initialize_ducklake(self):
        """ ATTACH to the ducklake. Ensure the needed sheets exist in the Excel file. 
        If they exist already, do an initial _pull. If not, do a _push """

        self.db.sql(f""" ATTACH 'ducklake:{self.local_catalog}' AS {self.ducklake_name} (DATA_PATH '{self.data_path}') """)

        tables = self.db.sql(f""" SELECT table_name FROM information_schema.tables where table_catalog like '__ducklake_metadata_{self.ducklake_name}' """).fetchall()

        self.catalog_tables = [t[0] for t in tables] # keep a list of table names, to iterate over later
        self.catalog_tables_no_ducklake = [t.replace('ducklake_','',1) for t in self.catalog_tables] # Excel sheetnames limited to 31 characters. Use shortened version

        if self.client_type == 'onedrive':
            # list sheets in excel file:
            remote_sheetnames, session_id = self.loop.run_until_complete(self._get_existing_sheets())

            # loop over catalog_tables and create ones that don't exist already
            all_sheets_exist_already = True # flag whether or not we start with a pull (if True) or a push (if False)
            missing_tables = []
            for t in self.catalog_tables_no_ducklake:
                if t in remote_sheetnames:
                    continue
                else:
                    all_sheets_exist_already = False
                    missing_tables.append(t)
            
            if len(missing_tables) > 0:
                # Create missing sheets concurrently
                self.loop.run_until_complete(self._create_sheets(missing_tables,session_id))

            if all_sheets_exist_already:
                # initialize state from the remote catalog
                # for now, this will error if the sheets exist but no column headers in sheet
                self._pull()
            else:
                # reset the remote to match the state of the local
                self._push()

            # close the workbook session
            self.loop.run_until_complete(self._close_workbook_session(session_id))

        elif self.client_type == 'excel':
            # check if all sheets are present that need to be:
            try:
                with pd.ExcelFile(self.excel_path) as f:
                    remote_sheetnames = f.sheet_names
            except FileNotFoundError:
                print(f'File {self.excel_path} not found. Attempting to create...')
                pd.DataFrame().to_excel(self.excel_path)
                print(f'File {self.excel_path} created!')
                remote_sheetnames = []

            all_sheets_exist_already = True # flag whether or not we start with a pull (if True) or a push (if False)
            for t in self.catalog_tables_no_ducklake:
                if t in remote_sheetnames:
                    continue
                else:
                    all_sheets_exist_already = False
            
            if all_sheets_exist_already:
                # initialize state from the remote catalog
                # for now, this will error if the sheets exist but no column headers in sheet
                self._pull()
            else:
                # reset the remote to match the state of the local
                self._push()

        else:
            raise NotImplementedError("Only OneDrive client type is currently supported")


    async def _create_workbook_session(self, persist_changes: bool) -> str:
        """Create a session id to tag a set of concurrent API calls with"""
        headers = self._acquire_token()
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/createSession"
        body = {"persistChanges": persist_changes}
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            data = await self._request_with_retry(session, 'post', url, headers=headers, json=body, allow_401_retries=True)
        return data.get('id')


    async def _close_workbook_session(self, session_id: str) -> None:
        """Close the workbook session"""
        headers = self._acquire_token()
        headers.update({'workbook-session-id': session_id})
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/closeSession"
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            await self._request_with_retry(session, 'post', url, headers=headers)


    async def _request_with_retry(self, session: aiohttp.ClientSession, method: str, url: str, headers : dict = None, allow_401_retries: bool = False, **kwargs) -> dict:
        """Wrapper to handle 429 and respect Retry-After header with exponential backoff for 404."""
        retry_for_404_limit = 5  # sometimes queries for ranges from the endpoint 404 if it hasn't caught up to the sheet creation
        retries_for_404 = 0
        retry_for_401_limit = 6  # may need a backoff period when first creating file for permissions to propagate
        retries_for_401 = 0
        while True:
            async with session.request(method, url, headers=headers, **kwargs) as resp:
                if resp.status == 404 and retries_for_404 < retry_for_404_limit:
                    retries_for_404 += 1
                    # exponential backoff with jitter
                    base_delay = 2 ** (retries_for_404 - 1)
                    jitter = np.random.uniform(0, 1)
                    delay = base_delay + jitter
                    print(f'retries_for_404 = {retries_for_404}, sleeping for {delay:.2f} seconds')
                    await asyncio.sleep(delay)
                elif (resp.status == 401) and (retries_for_401 < retry_for_401_limit) and allow_401_retries:
                    retries_for_401 += 1
                    # exponential backoff with jitter
                    base_delay = 2 ** (retries_for_401 - 1)
                    jitter = np.random.uniform(0, 1)
                    delay = base_delay + jitter
                    print(f'retries_for_401 = {retries_for_401}, sleeping for {delay:.2f} seconds')

                    # acquire a new token in case it needs to be updated to get permission on newly-created notebook
                    headers['Authorization'] = self._acquire_token(force_refresh=True)['Authorization']

                    await asyncio.sleep(delay)
                elif resp.status != 429:
                    resp.raise_for_status()
                    return await resp.json(content_type=None)  # sometimes we have no response body - ignore errors that would throw by setting content_type=None
                else:
                    retry_after = int(resp.headers.get('Retry-After', '1'))
                    await asyncio.sleep(retry_after)


    async def _get_existing_sheets(self) -> tuple:
        """Create a workbook session for initialization and fetch existing sheet names"""
        session_id = await self._create_workbook_session(persist_changes=True) # this call won't make changes, but if we have to create sheets, subsequent ones will reuse session_id
        headers = self._acquire_token()
        headers.update({'Workbook-Session-Id': session_id})

        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            response = await self._request_with_retry(
                session, 'get',
                f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets",
                headers=headers,
                allow_401_retries=True
            )
            sheet_names = [s['name'] for s in response.get('value', [])]

        return sheet_names, session_id
  

    async def _create_sheets(self, tables_to_create, session_id):
        headers = self._acquire_token()
        headers.update({'Workbook-Session-Id': session_id})
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            tasks = []
            for t in tables_to_create:
                add_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets/add"
                body = {"name": t}
                tasks.append(self._request_with_retry(session, 'post', add_url, headers=headers, json=body))
            await asyncio.gather(*tasks)

            
    def sql(self,query):
        """Run a DuckDB query. 
        Before the query, update the local catalog from the Excel. 
        After the query overwrite the remote Excel catalog
        Then return the original result"""

        self._pull()
        result = self.db.sql(query)
        self._push()

        return result
    

    def _pull(self):
        """ Iterate over all the catalog tables
        For each table, get the current values in the remote
        Accumulate the updates in a dict of table_name/dataframe, 
        and then truncate/overwrite all the local metadata tables
        Because everything coming from Excel may be a string, we also need to
        select 0 rows from the target table to grab the schema, so we can coerce the pandas df accordingly
        """
            
        # Determine dtypes for each table based on existing ducklake table schemas
        table_dtype_map = {}
        for t in self.catalog_tables_no_ducklake:
            t_dtypes = self.db.sql(f""" SELECT * FROM __ducklake_metadata_{self.ducklake_name}.{'ducklake_'+t} WHERE 1=0 """).df().dtypes
            # replace any int dtype with Int64 for nullable ints from pandas
            t_dtypes = t_dtypes.replace({
                'int32': 'Int64',
                'int64': 'Int64',
                })
            for col in t_dtypes.keys():
                # remove timezone localization, if any, imparted by pulling into pandas
                if pd.api.types.is_datetime64_any_dtype(t_dtypes[col]):
                    t_dtypes = t_dtypes.replace({t_dtypes[col]: 'datetime64[ns]'})
            table_dtype_map[t] = t_dtypes

        if self.client_type == 'onedrive':
            # Run async get for all tables
            metadata_to_write = self.loop.run_until_complete(self._session_pull_all(table_dtype_map))
        elif self.client_type == 'excel':
            metadata_to_write = {}
            with pd.ExcelFile(self.excel_path) as f:
                for t,dtypes in table_dtype_map.items():
                    # note we are iterating over sheets to pass a separate dtype dict for each sheet. May get small perf lift from reading all at once, if all cols with same name have same dtype

                    # if we leave 'bool' we error on blanks, since non-nullable. 
                    # if we make boolean, we need this fix, which looks like it doesn't 
                    # hit until version 3.0 of pandas: https://github.com/pandas-dev/pandas/pull/58994 
                    # hence this awkward workaround where we read as strings and coerce back to bool later
                    bool_cols = [k for k in dtypes.keys() if dtypes[k]=='bool']
                    dtypes = dtypes.replace({'bool' : 'object'}) 
                    metadata_to_write[t] = pd.read_excel(f,sheet_name=t,dtype=dict(dtypes))

                    # continuing our awkward boolean workaround:
                    for col in bool_cols:
                        list_of_bool_strings = metadata_to_write[t][col].to_list()
                        list_of_actual_bools = []
                        for b in list_of_bool_strings:
                            b = str(b) # in case it comes in as NaN
                            if b.lower().startswith('t'):
                                list_of_actual_bools.append(True)
                            elif b.lower().startswith('f'):
                                list_of_actual_bools.append(False)
                            else:
                                list_of_actual_bools.append(None)
                        metadata_to_write[t][col] = pd.Series(list_of_actual_bools,name=col,dtype=pd.BooleanDtype.name)

        else:
            raise NotImplementedError("Only OneDrive and Excel client types are currently supported")

        # write each table to the ducklake tables in DuckDB
        for t, df in metadata_to_write.items():
            this_table = f"__ducklake_metadata_{self.ducklake_name}.{'ducklake_'+t}"
            self.db.sql(f"""BEGIN TRANSACTION;
                            TRUNCATE {this_table};
                            INSERT INTO {this_table} SELECT * FROM df; 
                            COMMIT;""")


    async def _session_pull_all(self, table_dtype_map):
        """Create async task queue of all get requests for list of tables"""
        session_id = await self._create_workbook_session(persist_changes=False)
        headers = self._acquire_token()
        headers.update({'workbook-session-id': session_id})

        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            tasks = []
            for t, t_dtypes in table_dtype_map.items():
                tasks.append(self._async_pull_table(session, t, t_dtypes, headers)
            )
            results = await asyncio.gather(*tasks)

        await self._close_workbook_session(session_id)
        return dict(results)
    

    async def _async_pull_table(self, session, t, t_dtypes, headers) -> tuple:
        """ Get "used range" of sheet, which includes the values in the response.
         Then convert to a pandas dataframe """
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets('{t}')/usedRange"
        used_range = await self._request_with_retry(session, 'get', url, headers=headers)
        values = used_range.get('values', []) # will be a list of lists - each internal list is a row
        # convert to dict, then map to dataframe with correct types
        if len(values) > 1: # will be 1 with just header row
            keys = values[0]
            rows = values[1:]
            # Transpose to columns
            cols = list(zip(*rows))
            new_data_dict = {
                col: [None if v == '' else v for v in col_vals] # replace empty string '' with Python None, in case it goes in numeric column
                for col, col_vals in zip(keys, cols)
            }
            df = pd.DataFrame({col: pd.Series(new_data_dict[col], dtype=dt) for col, dt in t_dtypes.items()})
            # the ducklake_metadata table stores the encryption value as a string of 'true' or 'false
            # the roundtrip to Excel turns it into an Excel Boolean. since it's now a string again, need to 
            # convert to the expected case
            if t == 'metadata':
                df.loc[df.key == 'encrypted', 'value'] = df.loc[df.key == 'encrypted', 'value'].astype(str).str.lower()
        else:
            # create empty table with correct schema
            df = pd.DataFrame({col: pd.Series(dtype=dt) for col, dt in t_dtypes.items()})
        return (t, df)


    def _push(self):
        """ Iterate over all the catalog tables (concurrently with aiohttp)
        For each table, clear the current values in the remote (get used range and clear cells), 
        then replace with the full contents of the local
        """
        table_df_map = {}
        for t in self.catalog_tables_no_ducklake:
            t_df = self.db.sql(f""" SELECT * FROM __ducklake_metadata_{self.ducklake_name}.{'ducklake_'+t} """).df()
            for col in t_df.columns:
                # remove timezone localization, if any
                if pd.api.types.is_datetime64_any_dtype(t_df[col]):
                    if getattr(t_df[col].dt, 'tz', None) is not None:
                        t_df[col] = t_df[col].dt.tz_convert('UTC').dt.tz_localize(None)
            table_df_map[t] = t_df

        if self.client_type == 'onedrive':
            # Run async push for all tables
            self.loop.run_until_complete(self._session_push_all(table_df_map))
        elif self.client_type == 'excel':
            # TODO: mode 'a' doesn't work with fsspec files - need separate handling for excel via fsspec
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
                for t,t_df in table_df_map.items():
                    t_df.to_excel(writer, sheet_name=t, index=False)
        else:
            raise NotImplementedError("Only OneDrive and Excel client types are currently supported")


    async def _session_push_all(self, table_df_map):
        session_id = await self._create_workbook_session(persist_changes=True)
        headers = self._acquire_token()
        headers.update({'workbook-session-id': session_id})

        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            tasks = []
            for t, t_df in table_df_map.items():
                tasks.append(self._async_push_table(session, t, t_df, headers))
            await asyncio.gather(*tasks)

        await self._close_workbook_session(session_id)


    async def _async_push_table(self, session, t, t_df,headers):
        # Fetch used range to clear
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets('{t}')/usedRange"
        used_range = await self._request_with_retry(session, 'get', url, headers=headers)
        
        # clear used range of sheet
        clear_range = used_range['address'].split('!')[1] #splitting to get the range and not the sheetname
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets('{t}')/range(address='{clear_range}')/clear"
        body = {
            "apply_to": 'All'
        }
        clear_response = await self._request_with_retry(session,'post', url, headers=headers, json=body)

        # prep table values for patch request to update values in (cleared) range
        replace_map = {'<NA>': None, 'nan': None, 'None': None, 'NaT': None} # deal with the variety of ways a pandas NULL can be serialized to string
        values = [t_df.columns.tolist()] + t_df.astype(str).replace(replace_map).values.tolist()

        range_address = f"A1:{string.ascii_uppercase[t_df.shape[1]-1]}{t_df.shape[0]+1}"

        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/items/{self.item_id}/workbook/worksheets('{t}')/range(address='{range_address}')"

        # Request body with the values
        body = {
            "values": values
        }

        # Make the PATCH request to write data
        headers_patch = headers.copy()
        headers_patch['Content-Type'] = 'application/json' # add content type
        patch_resp = await self._request_with_retry(session, 'patch', url, headers=headers_patch, json=body)
        #print(f"Pushed {t}")         


    def _create_onedrive_workbook(self, fname: str, folder_path: str = None) -> None:
        """Create a new workbook on OneDrive with the given filename
        Since the Graph API doesn't currently support creating an empty xlsx file,
        create an empty xlsx locally, then upload it.

        folder path will be with respect to the root folder of the drive ID specified
        """

        # create a blank excel file
        wb = openpyxl.Workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)

        # upload using graph api
        with open(tmp_path, "rb") as f:
            response = self.loop.run_until_complete(self._async_create_file(fname,folder_path,f))

        if response:
            print(f"Workbook {fname} created successfully.")

            # the driveid + file ID becomes the excel_path
            self.item_id = response['id']

        else:
            print(f"Failed to create workbook {fname}: {response.text}")


    async def _async_create_file(self,fname,folder_path,f):
        headers = self._acquire_token()
        if folder_path:
            upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{folder_path.rstrip('/').lstrip('/')}/{fname}:/content"
        else:
            upload_url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{fname}:/content"
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=self.ssl_context, force_close=True)) as session:
            response = await self._request_with_retry(
                session, 'put',
                upload_url,
                headers=headers,
                data=f
            )

        return response


def test_onedrive():
    import time
    from datetime import datetime
    from dotenv import load_dotenv
    load_dotenv()
    MY_TEST_ONEDRIVE_PATH = os.getenv('MY_TEST_ONEDRIVE_PATH')
    print('creating test instance:')
    time_string = datetime.now().strftime('%Y-%m-%d_%H-%M-%S') 
    excel_fname = f'ducklake_test_{time_string}.xlsx'
    ducklake_file = f'ducklake_test_{time_string}.ducklake'
    start_time = time.time()
    test = DuckLakeXL(
        excel_path=excel_fname,
        data_path='../test/',
        ducklake_name='my_excel_ducklake',
        local_catalog=ducklake_file,
        drive_id=MY_TEST_ONEDRIVE_PATH.split('!')[0],
        folder_path='onedrive_ducklake',
        create_if_missing=True
    )
    print(f'Initialization took {time.time() - start_time:.2f} seconds')

    time.sleep(1.0)

    print('initialized...')
    start_time = time.time()
    test.sql("""USE my_excel_ducklake;
            CREATE TABLE my_table(id INTEGER, val VARCHAR);""")
    print(f'Table creation took {time.time() - start_time:.2f} seconds')

    time.sleep(1.0)

    print('table created...')
    start_time = time.time()
    test.sql("""INSERT INTO my_table VALUES
               (1, 'alpha'),
               (2, 'beta'),
                (3, 'gamma'),
                (4, 'delta');               
               """)
    print(f'Data insertion took {time.time() - start_time:.2f} seconds')

    time.sleep(1.0)

    print('data inserted...')
    start_time = time.time()
    test.sql("""SELECT * FROM my_table;""").show()
    print(f'First select took {time.time() - start_time:.2f} seconds')

    time.sleep(1.0)

    start_time = time.time()
    test.sql("""DELETE FROM my_table WHERE id = 3;""")
    print(f'Delete operation took {time.time() - start_time:.2f} seconds')

    time.sleep(1.0)

    start_time = time.time()
    test.sql("""SELECT * FROM my_table;""").show()
    print(f'Final select took {time.time() - start_time:.2f} seconds')

    print('OneDrive test complete!')


def test_excel():
    import time
    from datetime import datetime
    print('creating test instance:')
    time_string = datetime.now().strftime('%Y-%m-%d_%H-%M-%S') 
    excel_fname = f'../test/ducklake_test_{time_string}.xlsx'
    ducklake_file = f'ducklake_test_{time_string}.ducklake'
    start_time = time.time()
    test = DuckLakeXL(
        excel_path=excel_fname,
        data_path='../test/',
        ducklake_name='my_excel_ducklake',
        local_catalog=ducklake_file,
        drive_id=None,
        folder_path=None,
        create_if_missing=True
    )
    print(f'Initialization took {time.time() - start_time:.2f} seconds')

    print('initialized...')
    start_time = time.time()
    test.sql("""USE my_excel_ducklake;
            CREATE TABLE my_table(id INTEGER, val VARCHAR);""")
    print(f'Table creation took {time.time() - start_time:.2f} seconds')

    print('table created...')
    start_time = time.time()
    test.sql("""INSERT INTO my_table VALUES
               (1, 'alpha'),
               (2, 'beta'),
                (3, 'gamma'),
                (4, 'delta');               
               """)
    print(f'Data insertion took {time.time() - start_time:.2f} seconds')

    print('data inserted...')
    start_time = time.time()
    test.sql("""SELECT * FROM my_table;""").show()
    print(f'First select took {time.time() - start_time:.2f} seconds')

    start_time = time.time()
    test.sql("""DELETE FROM my_table WHERE id = 3;""")
    print(f'Delete operation took {time.time() - start_time:.2f} seconds')

    start_time = time.time()
    test.sql("""SELECT * FROM my_table;""").show()
    print(f'Final select took {time.time() - start_time:.2f} seconds')

    print('Excel test complete!')


def test_onedrive_existing_file():
    import time
    from datetime import datetime
    from dotenv import load_dotenv
    load_dotenv()
    MY_TEST_ONEDRIVE_PATH = os.getenv('MY_TEST_ONEDRIVE_PATH')
    print('creating test instance:')
    time_string = datetime.now().strftime('%Y-%m-%d_%H-%M-%S') 
    ducklake_file = f'ducklake_test_{time_string}.ducklake'
    start_time = time.time()
    test = DuckLakeXL(
        excel_path=MY_TEST_ONEDRIVE_PATH,
        data_path='../test/',
        ducklake_name='my_excel_ducklake',
        local_catalog=ducklake_file,
    )
    print(f'Initialization took {time.time() - start_time:.2f} seconds')


if __name__ == '__main__':
    test_excel()
    test_onedrive()
    test_onedrive_existing_file()