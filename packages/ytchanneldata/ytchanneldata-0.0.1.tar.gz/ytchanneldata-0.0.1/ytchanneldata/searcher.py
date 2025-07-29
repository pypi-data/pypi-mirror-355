import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from openpyxl import Workbook, load_workbook


def search(url):
    page = webdriver.Chrome()
    page.get(url)

    names = page.find_elements(By.XPATH, 
                                 "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-browse/div[4]/ytd-tabbed-page-header/tp-yt-app-header-layout/div/tp-yt-app-header/div[2]/div/div[2]/yt-page-header-renderer/yt-page-header-view-model/div/div[1]/div/yt-dynamic-text-view-model/h1/span")
    identifiers = page.find_elements(By.XPATH,
                               "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-browse/div[4]/ytd-tabbed-page-header/tp-yt-app-header-layout/div/tp-yt-app-header/div[2]/div/div[2]/yt-page-header-renderer/yt-page-header-view-model/div/div[1]/div/yt-content-metadata-view-model/div[1]/span/span")
    subs = page.find_elements(By.XPATH, 
                                     "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-browse/div[4]/ytd-tabbed-page-header/tp-yt-app-header-layout/div/tp-yt-app-header/div[2]/div/div[2]/yt-page-header-renderer/yt-page-header-view-model/div/div[1]/div/yt-content-metadata-view-model/div[2]/span[1]")
    videos = page.find_elements(By.XPATH, 
                                  "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-browse/div[4]/ytd-tabbed-page-header/tp-yt-app-header-layout/div/tp-yt-app-header/div[2]/div/div[2]/yt-page-header-renderer/yt-page-header-view-model/div/div[1]/div/yt-content-metadata-view-model/div[2]/span[3]")
    

    filename = 'YoutubeCHInfos.xlsx'

    if not os.path.exists(filename):
        workbook = Workbook()
        page_infos = workbook.active
        page_infos.title = 'Infos'
        page_infos = workbook['Infos']
    else:
        workbook = load_workbook(filename)
        page_infos = workbook['Infos']

    for name, identifier, sub, video in zip(names, identifiers, subs, videos):
        data = datetime.now().strftime("%d/%m/%Y")
        print(name.text)
        print(identifier.text)
        print(sub.text)
        print(video.text)
        page_infos.append([name.text, identifier.text, sub.text, video.text,data])

    workbook.save(filename)
    page.close()


