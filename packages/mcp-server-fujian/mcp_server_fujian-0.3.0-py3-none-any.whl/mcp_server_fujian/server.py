import os
import zipfile
import shutil
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from enum import Enum
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import ErrorData, Tool, TextContent, INVALID_PARAMS, INTERNAL_ERROR


class FujianTools(str, Enum):
    PROCESS_FUJIAN = "process_fujian"


def extract_zips(zip_file_path: str, output_dir: str) -> None:
    """解压 ZIP 文件及其嵌套 ZIP 文件到指定输出目录，并处理 Excel 文件格式。"""
    zip_path = Path(zip_file_path)
    output_dir = Path(output_dir)
    tmp_dir = output_dir / 'tmp'

    # 验证输入参数
    if not zip_path.exists() or not zip_path.is_file() or zip_path.suffix != '.zip':
        raise ErrorData(
            code=INVALID_PARAMS,
            message=f"无效的 ZIP 文件路径：{zip_file_path}，文件不存在或不是 ZIP 文件"
        )

    tmp_zip_dir = tmp_dir / f"zip_{zip_path.stem}"
    tmp_zip_dir.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(zip_path, 'r') as zfile:
            for zip_file in zfile.namelist():
                try:
                    # 处理文件名编码
                    decoded_name = zip_file.encode('cp437').decode('gbk', errors='replace')
                except Exception as e:
                    decoded_name = zip_file

                target_path = output_dir / decoded_name
                if target_path.exists():
                    continue

                # 解压到临时目录
                zfile.extract(zip_file, tmp_zip_dir)
                extracted_file_path = tmp_zip_dir / zip_file

                # 处理嵌套 ZIP 文件
                if decoded_name.endswith('.zip'):
                    try:
                        extract_zips(extracted_file_path, output_dir)
                    except ErrorData as ed:
                        raise ed
                    except Exception as e:
                        raise ErrorData(
                            code=INTERNAL_ERROR,
                            message=f"递归解压嵌套 ZIP 文件 {decoded_name} 失败：{str(e)}"
                        )
                    continue

                # 处理 Excel 文件
                os.makedirs(target_path.parent, exist_ok=True)
                if decoded_name.endswith(('.xlsx', '.xls')):
                    try:
                        wb = openpyxl.load_workbook(extracted_file_path)
                        wb.save(target_path)
                    except Exception as e:
                        os.rename(extracted_file_path, target_path)
                else:
                    # 其他文件直接移动
                    os.rename(extracted_file_path, target_path)


    except Exception as e:
        raise ErrorData(
            code=INTERNAL_ERROR,
            message=f"处理压缩包 {zip_path} 失败：{str(e)}"
        )
    finally:
        # 清理临时目录
        if tmp_zip_dir.exists():
            shutil.rmtree(tmp_zip_dir)


def check_and_extract(folder_path: str, keywords=None, exclude_keyword='芜湖路') -> tuple[dict, bool]:
    """筛选表格文件并分类为筛选表格、反馈文件和其他文件。"""
    if keywords is None:
        keywords = [
            '芜湖市供电公司', '镜湖区供电公司', '鸠江区供电公司',
            '弋江区供电公司', '湾沚区供电公司', '繁昌区供电公司',
            '南陵县供电公司', '无为市供电公司', '芜湖公司',
        ]

    extracted_files = {keyword: {'filtered': [], 'feedback': [], 'other': []} for keyword in keywords}
    has_non_feedback = False

    # 验证文件夹路径
    folder_path = Path(folder_path)
    if not folder_path.exists() or not folder_path.is_dir():
        raise ErrorData(
            code=INVALID_PARAMS,
            message=f"无效的文件夹路径：{folder_path}，目录不存在"
        )

    # 收集其他文件
    other_files = [
        os.path.join(root, file)
        for root, _, files in os.walk(folder_path)
        for file in files
        if not file.endswith(('.csv', '.xlsx', '.xls', '.zip'))
    ]

    # 处理表格文件
    for root, _, files in os.walk(folder_path):
        for file in files:
            if not file.endswith(('.csv', '.xlsx', '.xls')):
                continue

            file_path = os.path.join(root, file)

            try:
                if "反馈" not in file:
                    has_non_feedback = True
                    df = pd.read_excel(file_path) if file.endswith(('.xlsx', '.xls')) else pd.read_csv(file_path)
                    df = df.astype(str)

                    # 筛选数据
                    mask_include = df.apply(lambda x: x.str.contains('|'.join(keywords), na=False)).any(axis=1)
                    mask_exclude = df.apply(lambda x: x.str.contains(exclude_keyword, na=False)).any(axis=1)
                    mask = mask_include & ~mask_exclude

                    if mask.any():
                        extracted_data = df[mask]
                        for keyword in keywords:
                            keyword_mask = extracted_data.apply(lambda x: x.str.contains(keyword, na=False)).any(axis=1)
                            if not keyword_mask.any():
                                continue
                            new_file_path = os.path.join(root, f'extracted_{keyword}_{file}')
                            if file.endswith('.csv'):
                                extracted_data[keyword_mask].to_csv(new_file_path, index=False, encoding='utf-8-sig')
                            else:
                                wb = openpyxl.Workbook()
                                ws = wb.active
                                for col_num, column_title in enumerate(extracted_data[keyword_mask].columns, 1):
                                    ws.cell(row=1, column=col_num).value = column_title
                                for row_num, row_data in enumerate(extracted_data[keyword_mask].values, 2):
                                    for col_num, cell_value in enumerate(row_data, 1):
                                        ws.cell(row=row_num, column=col_num).value = cell_value
                                wb.save(new_file_path)
                            extracted_files[keyword]['filtered'].append(new_file_path)
                else:
                    for keyword in keywords:
                        extracted_files[keyword]['feedback'].append(file_path)
            except Exception as e:
                print(f"处理文件 {file_path} 失败：{str(e)}")
                continue

    for keyword in keywords:
        extracted_files[keyword]['other'] = other_files

    return extracted_files, has_non_feedback


def zip_files(folder_path: str, extracted_files: dict, has_non_feedback: bool) -> None:
    """根据关键词生成 ZIP 文件，或在无筛选表格时生成统一 ZIP 文件。"""
    folder_path = Path(folder_path)
    if not folder_path.exists():
        raise ErrorData(
            code=INVALID_PARAMS,
            message=f"无效的文件夹路径：{folder_path}，目录不存在"
        )

    if has_non_feedback:
        for keyword, file_dict in extracted_files.items():
            if not file_dict['filtered']:
                continue
            files = file_dict['filtered'] + file_dict['feedback'] + file_dict['other']
            if not files:
                continue
            zip_file_path = folder_path / f'extracted_files_{keyword}.zip'
            try:
                with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file in files:
                        zipf.write(file, os.path.relpath(file, folder_path))
            except Exception as e:
                raise ErrorData(
                    code=INTERNAL_ERROR,
                    message=f"生成压缩文件 {zip_file_path} 失败：{str(e)}"
                )
    else:
        feedback_files = list(set(extracted_files[list(extracted_files.keys())[0]]['feedback']))
        other_files = list(set(extracted_files[list(extracted_files.keys())[0]]['other']))
        files = feedback_files + other_files
        if not files:
            return
        zip_file_path = folder_path / 'all_feedback_and_others.zip'
        try:
            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in files:
                    zipf.write(file, os.path.relpath(file, folder_path))
        except Exception as e:
            raise ErrorData(
                code=INTERNAL_ERROR,
                message=f"生成统一压缩文件 {zip_file_path} 失败：{str(e)}"
            )
def process_files(self, zip_file_path: str, output_dir: str) -> str:
        """处理 ZIP 文件，解压、拆分表格并生成新的 ZIP 文件。"""
        try:
            extract_zips(zip_file_path, output_dir)
            extracted_files, has_non_feedback = check_and_extract(output_dir)
            zip_files(output_dir, extracted_files, has_non_feedback)
            return "文件处理成功"
        except ErrorData as ed:
            raise ed
        except Exception as e:
            raise ErrorData(
                code=INTERNAL_ERROR,
                message=f"处理过程中发生未知错误：{str(e)}"
            )
        finally:
            # 清理临时文件夹
            tmp_dir = Path(output_dir) / 'tmp'
            if tmp_dir.exists():
                shutil.rmtree(tmp_dir)


class FujianServer:
    # def process_info_codes(result: dict) -> None:
    #     """处理数据库查询结果中的 info_codes，调用 process_files 进行文件处理"""
    #     if 'results' in result:
    #         info_codes = [item['info_code'] for item in result['results']]
            
    #         for info_code in info_codes:
    #             zip_file_path = f"E:/芜湖/gzrw_file/gzrw_file/{info_code}.zip"
    #             output_dir = f"E:/芜湖/gzrw_file/{info_code}"
    #             if os.path.exists(zip_file_path):
    #                 process_files(zip_file_path, output_dir)
    def process_info_codes(info_code: str) -> None:
        """处理数据库查询结果中的 info_codes，调用 process_files 进行文件处理"""
        
        zip_file_path = f"E:/芜湖/gzrw_file/gzrw_file/{info_code}.zip"
        output_dir = f"E:/芜湖/gzrw_file/{info_code}"

        if os.path.exists(zip_file_path):
            process_files(zip_file_path, output_dir)
        else:
                print(f"压缩包未找到：{zip_file_path}")

async def serve() -> None:
    server = Server("mcp-fujian")
    fujian_server = FujianServer()

    @server.list_tools()
    async def list_tools() -> list[Tool]:
        return [
            Tool(
                name=FujianTools.PROCESS_FUJIAN.value,
                description="根据数据编号对指定的附件内容进行解析，包括表格拆分和压缩。",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "info_code": {
                            "type": "str",
                            "description": "数据编号",
                        },
                    },
                    "required": ["info_code"]
                },
            ),
        ]

    @server.call_tool()
    async def call_tool(name: str, arguments: dict) -> list[TextContent]:
        try:
            if name != FujianTools.PROCESS_FUJIAN.value:
                raise ErrorData(
                    code=INVALID_PARAMS,
                    message=f"未知工具：{name}"
                )

            if not all(k in arguments for k in ["info_code"]):
                raise ErrorData(
                    code=INVALID_PARAMS,
                    message="缺少必要的参数：info_code"
                )

            result = fujian_server.process_info_codes(arguments["info_code"])
            return [TextContent(type="text", text=result)]

        except ErrorData as ed:
            return [TextContent(type="error", text=f"错误：{ed.message}（错误代码：{ed.code}）")]
        except Exception as e:
            return [TextContent(
                type="error",
                text=f"处理 mcp-server-fujian 请求时发生未知错误：{str(e)}（错误代码：{INTERNAL_ERROR}）"
            )]

    options = server.create_initialization_options()
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, options, raise_exceptions=True)