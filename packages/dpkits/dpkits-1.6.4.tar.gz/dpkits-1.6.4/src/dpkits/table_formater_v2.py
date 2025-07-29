import multiprocessing
import xlwings as xw
import re as regex
import time
import datetime
from openpyxl.utils import get_column_letter
from colorama import Fore
from xlwings import sheets, books, Range
from .logging import Logging



class TableFormatterV2(Logging):

    def __init__(self, xlsx_name):
        super().__init__()
        self.xlsx_name = xlsx_name.rsplit('/', 1)[-1] if '/' in xlsx_name else xlsx_name




    def format_workbook(self, *, is_visible: bool = False):

        # is_visible = True

        st = time.time()

        app = xw.App(visible=is_visible)
        app.display_alerts = False

        try:

            wb = xw.Book(self.xlsx_name)

            for ws in wb.sheets:

                ws.activate()

                self.format_worksheet(wb=wb, ws=ws)


            str_output_name = str(self.xlsx_name).replace('.xlsx', f" - {datetime.datetime.now().strftime("%b %d")}.xlsx")


            if 'Content' in wb.sheet_names:
                wb.sheets['Content'].activate()

            wb.save(str_output_name)
            wb.close()

            app.display_alerts = True

        except Exception as ex:
            raise ex

        finally:
            app.quit()

        self.print(f">>> Sheet(s) formatting have completed in {datetime.timedelta(seconds=time.time() - st)}", self.clr_succ)


    def format_worksheet(self, wb: books, ws: sheets):

        if ws.name == 'Content':
            self.format_ws_content(wb=wb, ws=ws)
        else:
            self.format_ws_table(ws=ws)



    def format_ws_content(self, wb: books, ws: sheets):

        ws.range('B:B').column_width = 50

        lst_ws = wb.sheet_names

        for i, v in enumerate(lst_ws):

            if i == 0:
                continue

            content_cell = ws.range(f'B{i + 1}')
            content_cell.api.Hyperlinks.Add(Anchor=content_cell.api, Address=f"#'{v}'!B3", TextToDisplay=v)

            ws_tbl = wb.sheets[v]
            table_cell = ws_tbl.range('B3')
            table_cell.api.Hyperlinks.Add(Anchor=table_cell.api, Address=f"#'Content'!B{i + 1}", TextToDisplay='Content')

        self.print(['Format sheet', ws.name, 'Completed'], [None, self.clr_blue, None], sep=' | ')



    @staticmethod
    def hex_to_rgb_int(hex_color: str) -> int:
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        return rgb[2] * 65536 + rgb[1] * 256 + rgb[0]  # Convert to Excel RGB integer format



    @staticmethod
    def get_filter_range(visible_cells: Range) -> list:
        return regex.split(r'R|C', visible_cells.GetAddress(ReferenceStyle=False).split(',')[-1].replace(':', ''))[1:]



    def format_ws_table(self, ws: sheets):

        st = time.time()

        ws.range('B:B').column_width = 12
        ws.range('C:C').column_width = 30
        ws.range('F:F').column_width = 40

        first_row = ws.range('A2').end('down').row
        first_column = ws.range('A2').end('right').column

        last_cell = ws.used_range.last_cell
        last_row = last_cell.row
        last_column = last_cell.column

        is_sig_tbl = True

        if ws.range('C4').value is None or 'Weighted' in ws.range('C4').value:
            is_sig_tbl = False

            lst_col_remove = list(range(first_column + 1, last_column + 1, 2))
            lst_col_remove.reverse()

            for icol in lst_col_remove:
                ws.range(f'{get_column_letter(icol)}:{get_column_letter(icol)}').api.Delete()

            # ws.range(f'{first_row - 2}:{first_row - 1}').api.Delete()

            first_row = ws.range('A2').end('down').row
            first_column = ws.range('A2').end('right').column

            last_cell = ws.used_range.last_cell
            last_row = last_cell.row
            last_column = last_cell.column

        rng_content = ws.range(f'{get_column_letter(first_column)}{first_row}:{get_column_letter(last_column)}{last_row}')

        if 'percentage' in ws.range('C3').value:

            str_decimal = regex.findall(r'\(.+\)', ws.range('C3').value)[0][1:-1]


            rng_content.api.NumberFormat = str_decimal
        else:
            rng_content.api.NumberFormat = '0'


        idx_col_data = list(dict.fromkeys(ws.range(f'A2:A{last_row}').value))
        idx_col_data.remove(None)


        # BORDERING SIDE QRE--------------------------------------------------------------------------------------------
        for i in idx_col_data:
            ws.range('1:1').api.AutoFilter(Field=1, Criteria1=i)
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            lst_address = self.get_filter_range(visible_cells)
            self.format_ws_qre(ws=ws, lst_address=lst_address)
            ws.api.AutoFilterMode = False


        rng_all_qre_full = ws.range(f'B{first_row}:F{last_row}')
        rng_all_qre_full.api.Borders(11).Weight = 2

        # FORMAT ATTS: mean, std,...------------------------------------------------------------------------------------
        self.format_ws_qre_atts(ws=ws)


        # BORDERING HEADER----------------------------------------------------------------------------------------------
        dict_header_info = {
            'is_sig_tbl': is_sig_tbl,
            'first_row': 3,
            'first_col': first_column,
            'last_row': first_row - 1,
            'last_col': last_column,
            'color': {
                0: {'fontColor': '#FFFFFF', 'fgColor': '#153D64'},
                1: {'fontColor': '#FFFFFF', 'fgColor': '#215C98'},
                2: {'fontColor': '#000000', 'fgColor': '#4D93D9'},
                3: {'fontColor': '#000000', 'fgColor': '#A6C9EC'},

                4: {'fontColor': '#FFFFFF', 'fgColor': '#196B24'},
                5: {'fontColor': '#000000', 'fgColor': '#47D359'},
                6: {'fontColor': '#000000', 'fgColor': '#83E28E'},
                7: {'fontColor': '#000000', 'fgColor': '#C1F0C8'},

                8: {'fontColor': '#FFFFFF', 'fgColor': '#E97132'},
                9: {'fontColor': '#000000', 'fgColor': '#F1A983'},
                10: {'fontColor': '#000000', 'fgColor': '#F7C7AC'},
                11: {'fontColor': '#000000', 'fgColor': '#FBE2D5'},
            }
        }

        self.format_ws_header(ws=ws, dict_header_info=dict_header_info)

        self.print(['Format sheet', ws.name, f'Completed in {datetime.timedelta(seconds=round(time.time() - st, 0))}'], [None, self.clr_blue, None], sep=' | ')



    def format_ws_qre(self, *, ws: sheets, lst_address: list):

        start_row = int(lst_address[0])
        last_row = int(lst_address[2])

        start_col = int(lst_address[1])
        last_col = int(lst_address[-1])

        rng_base = ws.range(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{start_row}")
        rng_qre = ws.range(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}")

        self.print(['Format sheet', ws.name, rng_qre.address, f'{rng_base.value[1]}-{rng_base.value[3]}'], [None, self.clr_blue, self.clr_cyan_light, self.clr_magenta_light], sep=' | ', end='')

        rng_base.api.Font.Bold = True
        rng_base.api.Interior.Color = self.hex_to_rgb_int('#DAEEF3')
        rng_base.api.NumberFormat = '0'

        rng_qre.api.Borders.Weight = 2
        rng_qre.api.Borders(11).Weight = 1
        rng_qre.api.Borders(12).Weight = 1

        for icol in [1, 2]:

            rng_qre_name = ws.range(f"{get_column_letter(start_col + icol)}{start_row}:{get_column_letter(start_col + icol)}{last_row}")
            rng_qre_name.api.Merge()
            rng_qre_name.api.HorizontalAlignment = -4108
            rng_qre_name.api.VerticalAlignment = -4108
            rng_qre_name.api.WrapText = True

            if icol == 1:
                rng_qre_name.api.Font.Color = self.hex_to_rgb_int('#FFFFFF')

            rng_qre_name.api.Interior.Color = self.hex_to_rgb_int('#538DD5' if icol == 1 else '#D9E1F2')

        print(end='\r')


    def format_ws_qre_atts(self, *, ws: sheets):

        lst_att = [
            'mean', 'std', 'sem',
            'min', 'max',
            '25%', '50%', '75%',
            'calculate', 'friedman_pval',
            'combine', 'net',
            'net0', 'net1', 'net2',
        ]

        def format_att_font(*, rng: Range, attribute: str, is_entire_row: bool):

            match attribute:
                case str(x) if 'mean' in x or 'std' in x or 'sem' in x:

                    rng.Font.Bold = True

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                    else:
                        rng.Font.Color = self.hex_to_rgb_int('#0070C0')

                case str(x) if 'min' in x or 'max' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#00B050')

                case str(x) if '25%' in x or '50%' in x or '75%' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#963634')

                case str(x) if 'calculate' in x or 'friedman_pval' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#E26B0A')

                case str(x) if 'combine' in x:

                    if is_entire_row:
                        rng.Font.Bold = True

                    else:
                        rng.Font.Color = self.hex_to_rgb_int('#000000')


                case str(x) if 'net' in x:

                    if is_entire_row:
                        rng.Font.Bold = True

                    else:

                        match attribute:
                            case str(x) if 'net0' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#A02B93')
                                rng.Font.Color = self.hex_to_rgb_int('#FFFFFF')

                            case str(x) if 'net1' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#F2CEEF')

                            case str(x) if 'net2' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#B5E6A2')

                            case _:
                                rng.Font.Underline = 2  # Single underline
                                rng.Font.Color = self.hex_to_rgb_int('#000000')


        for att in lst_att:

            # FORMAT ENTIRE ROWS----------------------------------------------------------------------------------------
            ws.range(f'1:1').api.AutoFilter(Field=5, Criteria1=f'*{att}*')
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            format_att_font(rng=visible_cells, attribute=att, is_entire_row=True)
            ws.api.AutoFilterMode = False

            # FORMAT VAR_LABEL ONLY-------------------------------------------------------------------------------------
            ws.range(f'E1:F1').api.AutoFilter(Field=1, Criteria1=f'*{att}*')
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            format_att_font(rng=visible_cells, attribute=att, is_entire_row=False)
            ws.api.AutoFilterMode = False



    def format_ws_header(self, *, ws: sheets, dict_header_info: dict):

        is_sig_tbl, first_row, first_col, last_row, last_col, dict_color = dict_header_info.values()

        first_col_letter = get_column_letter(first_col)
        last_col_letter = get_column_letter(last_col)

        rng_full_header = ws.range(f"{first_col_letter}{first_row}:{last_col_letter}{last_row}")
        rng_full_header.api.Borders.Weight = 2

        for i, v in enumerate(range(first_row, last_row + 1)):
            rng_header_row = ws.range(f"{first_col_letter}{v}:{last_col_letter}{v}")
            rng_header_row.api.Interior.Color = self.hex_to_rgb_int(dict_color[i]['fgColor'])
            rng_header_row.api.Font.Color = self.hex_to_rgb_int(dict_color[i]['fontColor'])


        last_usage_row = ws.used_range.last_cell.row
        lst_format_row = list(range(first_row, last_row - 2))
        lst_format_row.reverse()
        lst_merge_range = list()  # [start_col, start_row, end_col, end_row, bold_num]

        bold_num = 2 if not is_sig_tbl else 3

        if is_sig_tbl:
            for icol in range(first_col, last_col, 2):
                lst_merge_range.append([icol, last_row, icol + 1, last_row, 0])
                lst_merge_range.append([icol, last_row - 2, icol + 1, last_row - 2, 0])


        for irow in lst_format_row:

            start_merge_col = first_col

            for icol in range(first_col + 1, last_col + 2):

                col_letter = get_column_letter(icol)
                prev_col_letter = get_column_letter(icol - 1)

                sig_char_cell = ws.range(f"{col_letter}{last_row}")

                prev_cell = ws.range(f"{prev_col_letter}{irow}")
                curr_cell = ws.range(f"{col_letter}{irow}")

                if irow == first_row:
                    upper_row_prev_cell = ws.range(f"{prev_col_letter}{irow}")
                    Upper_row_curr_cell = ws.range(f"{col_letter}{irow}")
                else:
                    upper_row_prev_cell = ws.range(f"{prev_col_letter}{irow - 1}")
                    Upper_row_curr_cell = ws.range(f"{col_letter}{irow - 1}")

                if (sig_char_cell.value == 'A' or icol == last_col + 1) and (prev_cell.value != curr_cell.value or upper_row_prev_cell.value != Upper_row_curr_cell.value):
                    lst_merge_range.append([start_merge_col, irow, icol - 1, irow, bold_num])
                    start_merge_col = icol



        def merge_by_range(lst_range: list):

            merge_range = ws.range(f"{get_column_letter(lst_range[0])}{lst_range[1]}:{get_column_letter(lst_range[2])}{lst_range[3]}")
            merge_range.api.Merge()
            merge_range.api.HorizontalAlignment = -4108  # Center alignment (xlCenter)
            merge_range.api.VerticalAlignment = -4108  # Center alignment (xlCenter)
            merge_range.api.Borders.Weight = 2

            header_x_side_range = ws.range(f"{get_column_letter(lst_range[0])}{lst_range[1]}:{get_column_letter(lst_range[2])}{last_usage_row}")

            if is_sig_tbl:
                if lst_range[-1] > 0:
                    header_x_side_range.api.Borders(7).Weight = lst_range[-1]
                    header_x_side_range.api.Borders(10).Weight = lst_range[-1]
                else:
                    header_x_side_range.api.Borders(7).Weight = 2
                    header_x_side_range.api.Borders(10).Weight = 2
                    header_x_side_range.api.Borders(11).Weight = 1

                    sig_col_range = ws.range(f"{get_column_letter(lst_range[2])}{lst_range[1]}:{get_column_letter(lst_range[2])}{last_usage_row}")
                    sig_col_range.api.Font.Bold = True
                    sig_col_range.api.Font.Color = self.hex_to_rgb_int('#FF0000')

            else:
                header_x_side_range.api.Borders(7).Weight = lst_range[-1]
                header_x_side_range.api.Borders(10).Weight = lst_range[-1]


        for item in lst_merge_range:
            merge_by_range(item)


        rng_full_header.api.Font.Bold = True
        rng_full_header.api.WrapText = True
        rng_full_header.api.HorizontalAlignment = -4108  # Center alignment (xlCenter)
        rng_full_header.api.VerticalAlignment = -4108  # Center alignment (xlCenter)


        # Freeze panes at the selected range

        freeze_range = ws.range(f'{get_column_letter(first_col)}{last_row + 1}')
        freeze_range.api.Application.ActiveWindow.FreezePanes = False
        freeze_range.api.Select()
        freeze_range.api.Application.ActiveWindow.FreezePanes = True


        if is_sig_tbl:
            ws.range(f'{last_row - 1}:{last_row - 1}').api.Delete()
        else:
            ws.range(f'{last_row - 1}:{last_row}').api.Delete()

        ws.range('D:E').api.EntireColumn.Hidden = True
        ws.range('1:2').api.EntireRow.Hidden = True
        ws.range('A:A').api.Delete()































