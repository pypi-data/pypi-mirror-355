import time
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import workbook, worksheet
from colorama import Fore
from .logging import Logging
import multiprocessing
import xlwings as xw
from xlwings import sheets, books, Range
import re as regex
from datetime import datetime



class TableFormatter(Logging):

    def __init__(self, xlsx_name):

        super().__init__()
        self.xlsx_name = xlsx_name.rsplit('/', 1)[-1] if '/' in xlsx_name else xlsx_name

        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.dot_border = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))

        self.dot_thin_right_border = Border(left=Side(style='dotted'), right=Side(style='thin'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        self.dot_thin_bot_border = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='thin'))
        self.dot_right_thin_right_border = Border(left=Side(style='dotted'), right=Side(style='thin'), top=Side(style='dotted'), bottom=Side(style='thin'))

        self.medium_left_1_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        self.medium_left_2_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='thin'))
        self.medium_left_3_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='thin'), bottom=Side(style='dotted'))
        self.medium_left_4_border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



    @staticmethod
    def format_content_sheet(ws_content, ws_dtbl, str_cell_content_addr):

        idx = ws_content.max_row + 1

        ws_dtbl[str_cell_content_addr].value = 'Content'
        ws_dtbl[str_cell_content_addr].hyperlink = f"#'Content'!B{idx}"
        ws_dtbl[str_cell_content_addr].font = Font(color='0000FF')

        ws_content[f'A{idx}'].value = ws_content.max_row
        ws_content[f'B{idx}'].value = ws_dtbl.title
        ws_content[f'B{idx}'].hyperlink = f"#'{ws_dtbl.title}'!{str_cell_content_addr}"
        ws_content[f'B{idx}'].font = Font(color='0000FF')



    def format_side_axis(self, ws, is_matrix_table, start_side_row, num_format, step, lst_sub_header_col, is_tbl_sig=False):

        lst_sub_side_col = list()

        start_row = start_side_row

        for irow in range(start_row, ws.max_row + 1):

            cur1_cell = ws.cell(irow, 1)

            cur1_cell.border = self.thin_border
            cur1_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cur1_cell.font = Font(bold=True, color='FFFFFF')
            cur1_cell.fill = PatternFill(patternType='solid', fgColor='538DD5')

            cur2_cell = ws.cell(irow, 2)
            cur2_cell.border = self.thin_border
            cur2_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cur2_cell.font = Font(bold=True)
            cur2_cell.fill = PatternFill(patternType='solid', fgColor='D9E1F2')

            if cur2_cell.value != ws.cell(irow + 1, 2).value or cur1_cell.value != ws.cell(irow + 1, 1).value:

                if '_Group' in str(ws.cell(irow + 1, 2).value) or '_Mean' in str(ws.cell(irow + 1, 2).value):
                    continue

                end_row = irow
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                start_row = irow + 1
                lst_sub_side_col.append(irow)

        if is_matrix_table:
            print(f"Formatting {ws.title} - Brand label")

            for icol in range(5, ws.max_column):
                for row in range(5, 7):
                    cur_cell = ws.cell(row, icol)
                    cur_cell.border = self.thin_border
                    cur_cell.font = Font(bold=True)

                    if row == 5:
                        cur_cell.alignment = Alignment(horizontal='center', vertical='center')
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='D8E4BC')
                    elif row == 6:
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='DAEEF3')

        start_row = start_side_row + 1 if is_matrix_table else start_side_row

        for irow in range(start_row, ws.max_row):

            if irow % 500 == 0:
                print(f"Formatting {ws.title} - row {irow}")

            if 'FT' in ws.cell(irow, 3).value:
                ws.row_dimensions[irow].hidden = True
            else:
                for icol in range(5, ws.max_column):

                    cur_cell = ws.cell(irow, icol)

                    cur_cell.border = self.dot_border

                    if icol in lst_sub_header_col:
                        cur_cell.border = self.dot_thin_right_border

                    if irow in lst_sub_side_col:
                        cur_cell.border = self.dot_thin_bot_border

                    if icol in lst_sub_header_col and irow in lst_sub_side_col:
                        cur_cell.border = self.dot_right_thin_right_border


                    if is_tbl_sig and ws.cell(start_side_row - 1, icol).value == 'A':
                        cur_cell.border = self.medium_left_1_border

                        if irow in lst_sub_side_col:
                            cur_cell.border = self.medium_left_2_border

                        if icol in lst_sub_header_col and irow in lst_sub_side_col:
                            cur_cell.border = self.medium_left_3_border

                        for irow_hd in range(3, start_side_row):
                            ws.cell(irow_hd, icol).border = self.medium_left_4_border

                    if ws.cell(irow, 4).value == 'base':
                        cur_cell.font = Font(bold=True)
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='DAEEF3')

                    elif ws.cell(irow, 4).value == 'bes':
                        cur_cell.font = Font(bold=True)
                        cur_cell.number_format = '0.0'
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='FDE9D9')

                        if icol == 6 or (icol > 6 and (icol - 6) % step == 0):
                            cur_cell.alignment = Alignment(horizontal='center', vertical='center')
                            ws.merge_cells(start_row=irow, start_column=icol, end_row=irow, end_column=icol + step - 1)

                    else:
                        if ws.cell(irow, 4).value in ['mean', 'std']:
                            ws.cell(irow, 5).font = Font(bold=True, color='0070C0')
                            cur_cell.font = Font(bold=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 4).value in ['min', 'max']:
                            ws.cell(irow, 5).font = Font(bold=True, color='00B050')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 4).value in ['25%', '50%', '75%']:
                            ws.cell(irow, 5).font = Font(bold=True, color='963634')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif 'calculate' in ws.cell(irow, 4).value:
                            ws.cell(irow, 5).font = Font(bold=True, color='E26B0A')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif 'friedman_pval' in ws.cell(irow, 4).value:
                            ws.cell(irow, 5).font = Font(bold=True, color='E26B0A')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 3).value == 'GROUP':
                            cur_cell.font = Font(bold=True)

                            if num_format == 'pct':
                                cur_cell.number_format = '0'
                            elif num_format == 'pct_sign':
                                cur_cell.number_format = '0%'

                        elif int(ws.cell(irow, 4).value) > 90000:

                            cur_cell.font = Font(bold=True)

                            if num_format == 'pct':
                                cur_cell.number_format = '0'
                            elif num_format == 'pct_sign':
                                cur_cell.number_format = '0%'

                            cat_lbl = str(ws.cell(irow, 5).value)
                            net_cell = ws.cell(irow, icol)

                            # NET CODE UPDATE FOR MSN PROJECTS
                            if 'NET 0' in cat_lbl.upper():
                                # Net: #A02B93
                                net_cell.fill = PatternFill(patternType='solid', fgColor='A02B93')
                                net_cell.font = Font(bold=True, color='FFFFFF')

                            elif 'NET 1' in cat_lbl.upper():
                                # Sub-net: #F2CEEF
                                net_cell.fill = PatternFill(patternType='solid', fgColor='F2CEEF')

                            elif 'NET 2' in cat_lbl.upper():
                                # Sub-sub-net: #B5E6A2
                                net_cell.fill = PatternFill(patternType='solid', fgColor='B5E6A2')

                            else:
                                if '(NET)' in cat_lbl.upper():
                                    ws.cell(irow, 5).font = Font(bold=True, underline='double')
                                else:
                                    ws.cell(irow, 5).font = Font(bold=True, underline='single')

                        else:
                            if icol > 5:
                                if num_format == 'pct':
                                    cur_cell.number_format = '0'
                                elif num_format == 'pct_sign':
                                    cur_cell.number_format = '0%'

                        if is_tbl_sig and icol % 2 != 0 and icol > 5:
                            cur_cell.font = Font(bold=True, color='FF0000')



    def format_sig_table(self):

        wb = openpyxl.load_workbook(self.xlsx_name, data_only=True)

        ws_content = wb['Content']
        ws_content.column_dimensions['B'].width = 60

        for ws in wb.worksheets:

            if ws.title in ['Content']:
                continue

            print(f"Formatting {ws.title}")

            self.format_content_sheet(ws_content, ws, str_cell_content_addr='A3')

            lst_sub_header_col = list()

            last_header_row = 2

            while ws.cell(last_header_row, 4).value != 'base':
                last_header_row += 1

                if last_header_row >= 30:
                    last_header_row = -999
                    break

            last_header_row -= 1

            if not ws.cell(4, 2).value:
                is_sig_tbl = False
            else:
                is_sig_tbl = True if 'Dependent' in ws.cell(4, 2).value or 'Independent' in ws.cell(4, 2).value else False

            if not is_sig_tbl:
                ws.delete_rows(last_header_row - 1, 2)
                last_header_row = last_header_row - 2
            else:
                ws.delete_rows(last_header_row - 1, 1)
                last_header_row = last_header_row - 1

            ws.freeze_panes = f'F{last_header_row + 1}'

            if 'count' in ws.cell(3, 2).value:
                num_format = 'count'
            else:
                if '%' in ws.cell(3, 2).value:
                    num_format = 'pct_sign'
                else:
                    num_format = 'pct'

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['E'].width = 40

            if not is_sig_tbl:

                for icol in range(6, ws.max_column + 1).__reversed__():
                    if icol % 2 != 0:
                        # ws.delete_cols(icol, 1)
                        ws.delete_cols(icol)

            else:
                for icol in range(6, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(icol)].width = 7


            ws.column_dimensions['C'].hidden = True
            ws.column_dimensions['D'].hidden = True

            ws.row_dimensions[1].hidden = True
            ws.row_dimensions[2].hidden = True

            if is_sig_tbl and last_header_row > 1:
                ws.row_dimensions[last_header_row - 1].height = 30
            else:
                ws.row_dimensions[last_header_row].height = 30

            print(f"Formatting {ws.title} - Header")

            ws_max_col = ws.max_column

            dict_header_color = {
                0: {'fontColor': 'FFFFFF', 'fgColor': '203764'},
                1: {'fontColor': 'FFFFFF', 'fgColor': '305496'},
                2: {'fontColor': '000000', 'fgColor': '8EA9DB'},
                3: {'fontColor': '000000', 'fgColor': 'B4C6E7'},
                4: {'fontColor': '000000', 'fgColor': 'D9E1F2'},

                5: {'fontColor': '000000', 'fgColor': 'B5F1CD'},
                6: {'fontColor': '000000', 'fgColor': '57DF8E'},
                7: {'fontColor': '000000', 'fgColor': '27CF6B'},
                8: {'fontColor': 'FFFFFF', 'fgColor': '1E9E52'},
                9: {'fontColor': 'FFFFFF', 'fgColor': '156e3a'},

                # 5: {'fontColor': '000000', 'fgColor': 'A9D08E'},
                # 6: {'fontColor': '000000', 'fgColor': 'C6E0B4'},
                # 7: {'fontColor': '000000', 'fgColor': 'E2EFDA'},
                # 8: {'fontColor': '000000', 'fgColor': 'F4B084'},
                # 9: {'fontColor': '000000', 'fgColor': 'F8CBAD'},

                10: {'fontColor': '000000', 'fgColor': 'FCE4D6'},
                11: {'fontColor': '000000', 'fgColor': 'FCD5B4'},
                12: {'fontColor': '000000', 'fgColor': 'FDE9D9'},
            }

            for irow in range(3, last_header_row + 1):
                start_column = 6

                for icol in range(6, ws_max_col + 1):

                    cur_cell = ws.cell(irow, icol)

                    cur_cell.border = self.thin_border
                    cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cur_cell.font = Font(bold=True, size=12, color=dict_header_color[irow - 3]['fontColor'])
                    cur_cell.fill = PatternFill(patternType='solid', fgColor=dict_header_color[irow - 3]['fgColor'])

                    if irow <= 4:

                        if cur_cell.value != ws.cell(irow, icol + 1).value:
                            end_column = icol

                            if start_column != end_column and cur_cell.value is not None:

                                ws.merge_cells(start_row=irow, start_column=start_column, end_row=irow, end_column=end_column)
                                lst_sub_header_col.append(icol)

                            start_column = icol + 1

                    else:

                        if cur_cell.value != ws.cell(irow, icol + 1).value or ws.cell(irow - 1, icol + 1).value is not None:
                            end_column = icol

                            if start_column != end_column and cur_cell.value is not None:

                                ws.merge_cells(start_row=irow, start_column=start_column, end_row=irow, end_column=end_column)
                                lst_sub_header_col.append(icol)

                            start_column = icol + 1


            print(f"Formatting {ws.title} - Side axis")



            start_side_row = last_header_row + 1
            self.format_side_axis(ws=ws, is_matrix_table=False, start_side_row=start_side_row, num_format=num_format,
                                  step=1, lst_sub_header_col=lst_sub_header_col, is_tbl_sig=is_sig_tbl)


        # output_name = self.xlsx_name.replace('.xlsx', '_output.xlsx')
        output_name = self.xlsx_name

        print(f"Save wb as {output_name}")
        wb.save(output_name)
        wb.close()



    """
    --------------------------------------------------------------------------------------------------------------------
    ADD NEW-------------------------------------------------------------------------------------------------------------
    --------------------------------------------------------------------------------------------------------------------
    """



    def format_workbook(self):

        is_visible = False

        app = xw.App(visible=is_visible)
        app.display_alerts = False

        try:

            wb = xw.Book(self.xlsx_name)

            for ws in wb.sheets:

                ws.activate()

                self.format_worksheet(wb=wb, ws=ws)


            str_output_name = str(self.xlsx_name).replace('.xlsx', f" - {datetime.now().strftime("%b %d")}.xlsx")

            wb.save(str_output_name)
            wb.close()

            app.display_alerts = True

        except Exception as e:
            raise e

        finally:
            app.quit()



    def format_worksheet(self, wb: books, ws: sheets):

        if ws.name == 'Content':
            self.format_ws_content(wb=wb, ws=ws)
        else:
            self.format_ws_table(ws=ws)




    def format_ws_content(self, wb: books, ws: sheets):
        self.print(f"Format worksheet {ws.name}")

        ws.range('B:B').column_width = 50

        lst_ws = wb.sheet_names

        for i, v in enumerate(lst_ws):

            if i == 0:
                continue

            content_cell = ws.range(f'B{i + 1}')
            content_cell.api.Hyperlinks.Add(Anchor=content_cell.api, Address=f"#'{v}'!B3", TextToDisplay=v)

            ws_tbl = wb.sheets[v]
            table_cell = ws_tbl.range('B3')
            table_cell.api.Hyperlinks.Add(Anchor=table_cell.api, Address=f"#'Content'!B{i + 1}", TextToDisplay='Content')



    @staticmethod
    def hex_to_rgb_int(hex_color: str) -> int:
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        return rgb[2] * 65536 + rgb[1] * 256 + rgb[0]  # Convert to Excel RGB integer format



    @staticmethod
    def get_filter_range(visible_cells: Range) -> list:
        return regex.split(r'R|C', visible_cells.GetAddress(ReferenceStyle=False).split(',')[-1].replace(':', ''))[1:]



    def format_ws_table(self, ws: sheets):

        self.print(f"Format worksheet {ws.name}")

        ws.range('B:B').column_width = 12
        ws.range('C:C').column_width = 30
        ws.range('F:F').column_width = 40

        first_row = ws.range('A2').end('down').row
        first_column = ws.range('A2').end('right').column

        last_cell = ws.used_range.last_cell
        last_row = last_cell.row
        last_column = last_cell.column

        is_sig_tbl = True

        if ws.range('C4').value is None:
            is_sig_tbl = False

            lst_col_remove = list(range(first_column + 1, last_column + 1, 2))
            lst_col_remove.reverse()

            for icol in lst_col_remove:
                ws.range(f'{get_column_letter(icol)}:{get_column_letter(icol)}').api.Delete()

            # ws.range(f'{first_row - 2}:{first_row - 1}').api.Delete()

            first_row = ws.range('A2').end('down').row
            first_column = ws.range('A2').end('right').column

            last_cell = ws.used_range.last_cell
            last_row = last_cell.row
            last_column = last_cell.column

        rng_content = ws.range(f'{get_column_letter(first_column)}{first_row}:{get_column_letter(last_column)}{last_row}')


        match ws.range('C3').value:
            case 'Cell content: percentage(%)':
                rng_content.api.NumberFormat = '0%'

            case _:
                rng_content.api.NumberFormat = '0'

        idx_col_data = list(dict.fromkeys(ws.range(f'A2:A{last_row}').value))
        idx_col_data.remove(None)


        # BORDERING SIDE QRE--------------------------------------------------------------------------------------------
        for i in idx_col_data:
            ws.range('1:1').api.AutoFilter(Field=1, Criteria1=i)
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            lst_address = self.get_filter_range(visible_cells)
            self.format_ws_qre(ws=ws, lst_address=lst_address)
            ws.api.AutoFilterMode = False


        rng_all_qre_full = ws.range(f'B{first_row}:F{last_row}')
        rng_all_qre_full.api.Borders(11).Weight = 2

        # FORMAT ATTS: mean, std,...------------------------------------------------------------------------------------
        self.format_ws_qre_atts(ws=ws)


        # BORDERING HEADER----------------------------------------------------------------------------------------------
        dict_header_info = {
            'is_sig_tbl': is_sig_tbl,
            'first_row': 3,
            'first_col': first_column,
            'last_row': first_row - 1,
            'last_col': last_column,
            'color': {
                0: {'fontColor': '#FFFFFF', 'fgColor': '#153D64'},
                1: {'fontColor': '#FFFFFF', 'fgColor': '#215C98'},
                2: {'fontColor': '#000000', 'fgColor': '#4D93D9'},
                3: {'fontColor': '#000000', 'fgColor': '#A6C9EC'},

                4: {'fontColor': '#FFFFFF', 'fgColor': '#196B24'},
                5: {'fontColor': '#000000', 'fgColor': '#47D359'},
                6: {'fontColor': '#000000', 'fgColor': '#83E28E'},
                7: {'fontColor': '#000000', 'fgColor': '#C1F0C8'},

                8: {'fontColor': '#FFFFFF', 'fgColor': '#E97132'},
                9: {'fontColor': '#000000', 'fgColor': '#F1A983'},
                10: {'fontColor': '#000000', 'fgColor': '#F7C7AC'},
                11: {'fontColor': '#000000', 'fgColor': '#FBE2D5'},
            }
        }

        self.format_ws_header(ws=ws, dict_header_info=dict_header_info)





    def format_ws_qre(self, *, ws: sheets, lst_address: list):

        start_row = int(lst_address[0])
        last_row = int(lst_address[2])

        start_col = int(lst_address[1])
        last_col = int(lst_address[-1])

        rng_base = ws.range(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{start_row}")
        rng_qre = ws.range(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}")

        rng_base.api.Font.Bold = True
        rng_base.api.Interior.Color = self.hex_to_rgb_int('#DAEEF3')
        rng_base.api.NumberFormat = '0'

        rng_qre.api.Borders.Weight = 2
        rng_qre.api.Borders(11).Weight = 1
        rng_qre.api.Borders(12).Weight = 1

        for icol in [1, 2]:

            rng_qre_name = ws.range(f"{get_column_letter(start_col + icol)}{start_row}:{get_column_letter(start_col + icol)}{last_row}")
            rng_qre_name.api.Merge()
            rng_qre_name.api.HorizontalAlignment = -4108
            rng_qre_name.api.VerticalAlignment = -4108
            rng_qre_name.api.WrapText = True

            if icol == 1:
                rng_qre_name.api.Font.Color = self.hex_to_rgb_int('#FFFFFF')

            rng_qre_name.api.Interior.Color = self.hex_to_rgb_int('#538DD5' if icol == 1 else '#D9E1F2')



    def format_ws_qre_atts(self, *, ws: sheets):

        lst_att = [
            'mean', 'std',
            'min', 'max',
            '25%', '50%', '75%',
            'calculate', 'friedman_pval',
            'combine', 'net',
            'net0', 'net1', 'net2',
        ]

        def format_att_font(*, rng: Range, attribute: str, is_entire_row: bool):

            match attribute:
                case str(x) if 'mean' in x or 'std' in x:

                    rng.Font.Bold = True

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                    else:
                        rng.Font.Color = self.hex_to_rgb_int('#0070C0')

                case str(x) if 'min' in x or 'max' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#00B050')

                case str(x) if '25%' in x or '50%' in x or '75%' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#963634')

                case str(x) if 'calculate' in x or 'friedman_pval' in x:

                    if is_entire_row:
                        rng.NumberFormat = '0.00'
                        rng.Font.Italic = True

                    else:
                        rng.Font.Bold = True
                        rng.Font.Color = self.hex_to_rgb_int('#E26B0A')

                case str(x) if 'combine' in x:

                    if is_entire_row:
                        rng.Font.Bold = True

                    else:
                        rng.Font.Color = self.hex_to_rgb_int('#000000')


                case str(x) if 'net' in x:

                    if is_entire_row:
                        rng.Font.Bold = True

                    else:

                        match attribute:
                            case str(x) if 'net0' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#A02B93')
                                rng.Font.Color = self.hex_to_rgb_int('#FFFFFF')

                            case str(x) if 'net1' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#F2CEEF')

                            case str(x) if 'net2' in x:
                                rng.Interior.Color = self.hex_to_rgb_int('#B5E6A2')

                            case _:
                                rng.Font.Underline = 2  # Single underline
                                rng.Font.Color = self.hex_to_rgb_int('#000000')


        for att in lst_att:

            # FORMAT ENTIRE ROWS----------------------------------------------------------------------------------------
            ws.range(f'1:1').api.AutoFilter(Field=5, Criteria1=f'*{att}*')
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            format_att_font(rng=visible_cells, attribute=att, is_entire_row=True)
            ws.api.AutoFilterMode = False

            # FORMAT VAR_LABEL ONLY-------------------------------------------------------------------------------------
            ws.range(f'E1:F1').api.AutoFilter(Field=1, Criteria1=f'*{att}*')
            filter_range = ws.api.AutoFilter.Range
            visible_cells = filter_range.SpecialCells(12)
            format_att_font(rng=visible_cells, attribute=att, is_entire_row=False)
            ws.api.AutoFilterMode = False



    def format_ws_header(self, *, ws: sheets, dict_header_info: dict):

        is_sig_tbl, first_row, first_col, last_row, last_col, dict_color = dict_header_info.values()

        first_col_letter = get_column_letter(first_col)
        last_col_letter = get_column_letter(last_col)

        rng_full_header = ws.range(f"{first_col_letter}{first_row}:{last_col_letter}{last_row}")
        rng_full_header.api.Borders.Weight = 2

        for i, v in enumerate(range(first_row, last_row + 1)):
            rng_header_row = ws.range(f"{first_col_letter}{v}:{last_col_letter}{v}")
            rng_header_row.api.Interior.Color = self.hex_to_rgb_int(dict_color[i]['fgColor'])
            rng_header_row.api.Font.Color = self.hex_to_rgb_int(dict_color[i]['fontColor'])


        last_usage_row = ws.used_range.last_cell.row
        lst_format_row = list(range(first_row, last_row - 2))
        lst_format_row.reverse()
        lst_merge_range = list()  # [start_col, start_row, end_col, end_row, bold_num]

        bold_num = 2 if not is_sig_tbl else 3

        if is_sig_tbl:
            for icol in range(first_col, last_col, 2):
                lst_merge_range.append([icol, last_row, icol + 1, last_row, 0])
                lst_merge_range.append([icol, last_row - 2, icol + 1, last_row - 2, 0])




        for irow in lst_format_row:

            start_merge_col = first_col

            for icol in range(first_col + 1, last_col + 2):

                col_letter = get_column_letter(icol)
                prev_col_letter = get_column_letter(icol - 1)

                sig_char_cell = ws.range(f"{col_letter}{last_row}")

                prev_cell = ws.range(f"{prev_col_letter}{irow}")
                curr_cell = ws.range(f"{col_letter}{irow}")

                if irow == first_row:
                    upper_row_prev_cell = ws.range(f"{prev_col_letter}{irow}")
                    Upper_row_curr_cell = ws.range(f"{col_letter}{irow}")
                else:
                    upper_row_prev_cell = ws.range(f"{prev_col_letter}{irow - 1}")
                    Upper_row_curr_cell = ws.range(f"{col_letter}{irow - 1}")

                if (sig_char_cell.value == 'A' or icol == last_col + 1) and (prev_cell.value != curr_cell.value or upper_row_prev_cell.value != Upper_row_curr_cell.value):
                    lst_merge_range.append([start_merge_col, irow, icol - 1, irow, bold_num])
                    start_merge_col = icol






        def merge_by_range(lst_range: list):

            merge_range = ws.range(f"{get_column_letter(lst_range[0])}{lst_range[1]}:{get_column_letter(lst_range[2])}{lst_range[3]}")
            merge_range.api.Merge()
            merge_range.api.HorizontalAlignment = -4108  # Center alignment (xlCenter)
            merge_range.api.VerticalAlignment = -4108  # Center alignment (xlCenter)
            merge_range.api.Borders.Weight = 2

            header_x_side_range = ws.range(f"{get_column_letter(lst_range[0])}{lst_range[1]}:{get_column_letter(lst_range[2])}{last_usage_row}")

            if is_sig_tbl:
                if lst_range[-1] > 0:
                    header_x_side_range.api.Borders(7).Weight = lst_range[-1]
                    header_x_side_range.api.Borders(10).Weight = lst_range[-1]
                else:
                    header_x_side_range.api.Borders(7).Weight = 2
                    header_x_side_range.api.Borders(10).Weight = 2
                    header_x_side_range.api.Borders(11).Weight = 1

                    sig_col_range = ws.range(f"{get_column_letter(lst_range[2])}{lst_range[1]}:{get_column_letter(lst_range[2])}{last_usage_row}")
                    sig_col_range.api.Font.Bold = True
                    sig_col_range.api.Font.Color = self.hex_to_rgb_int('#FF0000')

            else:
                header_x_side_range.api.Borders(7).Weight = lst_range[-1]
                header_x_side_range.api.Borders(10).Weight = lst_range[-1]


        for item in lst_merge_range:
            merge_by_range(item)


        rng_full_header.api.Font.Bold = True
        rng_full_header.api.WrapText = True
        rng_full_header.api.HorizontalAlignment = -4108  # Center alignment (xlCenter)
        rng_full_header.api.VerticalAlignment = -4108  # Center alignment (xlCenter)


        # Freeze panes at the selected range

        freeze_range = ws.range(f'{get_column_letter(first_col)}{last_row + 1}')
        freeze_range.api.Application.ActiveWindow.FreezePanes = False
        freeze_range.api.Select()
        freeze_range.api.Application.ActiveWindow.FreezePanes = True


        if is_sig_tbl:
            ws.range(f'{last_row - 1}:{last_row - 1}').api.Delete()
        else:
            ws.range(f'{last_row - 1}:{last_row}').api.Delete()

        ws.range('D:E').api.EntireColumn.Hidden = True
        ws.range('1:2').api.EntireRow.Hidden = True
        ws.range('A:A').api.Delete()































