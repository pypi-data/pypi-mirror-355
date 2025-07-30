"""
Some helper functions to work with openpyxl: write tables, use autofilters, read_excel when it fails...
"""
from ong_utils.import_utils import raise_extra_exception
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import TableStyleInfo, Table
except ModuleNotFoundError:
    raise_extra_exception("xlsx")


def df_to_excel(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name: str, index=False, add_table: bool = True):
    """
    Writes a df to an opened Excel, fitting column with and adding Tables or activating autofilter
    Args:
        df: DataFrame to write
        writer: already opened excel writer, with engine = "openpyxl" (otherwise a ValueError will be raised)
        sheet_name: name of the sheet to write in
        index: False (default) to not write index to excel
        add_table: True (default) to add Tables to excel, False to add just autofilter

    Returns:
        None
    """
    if writer.engine != "openpyxl":
        raise ValueError(f"Error: writer engine ({writer.engine}) must be openpyxl, "
                         f"e.g.: pd.ExcelWriter('filename.xlsx', engine='openpyxl')")
    cols = [c.upper() if isinstance(c, str) else c for c in df.columns]
    if len(set(cols)) < len(cols):  # There are duplicated values
        new_cols = [f"{c}_{i}" if c.upper() in cols[i + 1:] else c for i, c in enumerate(df.columns)]
        df.columns = new_cols

    if df.empty:
        df.loc[0] = None  # Force empty row
    df.to_excel(writer, sheet_name=sheet_name, index=index, header=True)
    ws = writer.sheets[sheet_name]
    columns = ws.columns
    for column, xls_column in zip(df, columns):
        column_length = max(df[column].astype(str).map(len).max(), len(str(column)))
        column_length = max(column_length, len(str(column)))  # Include headers in calculation
        ws.column_dimensions[xls_column[0].column_letter].width = column_length

    if add_table:
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # create a table
        tab = Table(displayName=sheet_name.replace(" ", "_"), ref=ws.dimensions, tableStyleInfo=style)
        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unique through out defined names and all other table name. 
        '''
        ws.add_table(tab)
    else:
        # Add just autofilter
        ws.auto_filter.ref = ws.dimensions


# Currently unused
def read_excel_file(filename: str, sheet_name: str) -> pd.DataFrame:
    """Reads Excel file from a filename. If it fails, uses directly openpyxl to process file"""
    try:
        retval = pd.read_excel(filename, sheet_name=sheet_name)
        return retval
        # with pd.ExcelFile(filename) as excel:
        #     retval = pd.read_excel(excel, sheet_name=sheet_name)
        # return retval
    except UnicodeError as ue:
        print(ue)
        # Use openpyxl directly to read the dataframe
        wb = load_workbook(filename=filename)
        sheet = wb.worksheets[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        retval = pd.DataFrame(data)
        return retval
